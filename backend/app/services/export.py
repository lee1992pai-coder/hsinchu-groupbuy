"""
揀貨單匯出：PDF 與 Excel
"""
import io
from datetime import datetime


def build_pick_list_excel(merchant_name: str, items: list[dict]) -> bytes:
    """
    items: [{"product": "肉乾原味", "variant": "大箱", "quantity": 12}, ...]
    回傳 xlsx bytes
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "揀貨總表"

    # 標題列
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{merchant_name} 揀貨單 — {datetime.now().strftime('%Y/%m/%d %H:%M')}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["項次", "商品名稱", "規格", "需備數量"]
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, item in enumerate(items, 1):
        ws.append([
            idx,
            item.get("product", ""),
            item.get("variant", "-"),
            item.get("quantity", 0),
        ])

    # 欄寬
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pick_list_pdf(merchant_name: str, items: list[dict]) -> bytes:
    """
    使用 reportlab 產生 PDF 揀貨單
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = styles["Title"]
    title_style.fontName = "STSong-Light"
    story.append(Paragraph(
        f"{merchant_name} 揀貨單 — {datetime.now().strftime('%Y/%m/%d %H:%M')}",
        title_style,
    ))
    story.append(Spacer(1, 12))

    table_data = [["項次", "商品名稱", "規格", "需備數量"]]
    for idx, item in enumerate(items, 1):
        table_data.append([
            str(idx),
            item.get("product", ""),
            item.get("variant", "-"),
            str(item.get("quantity", 0)),
        ])

    t = Table(table_data, colWidths=[40, 220, 100, 80])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(t)

    doc.build(story)
    return buf.getvalue()
